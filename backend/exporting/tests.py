import csv
import io

import pytest
from rest_framework.test import APIClient

from accounts.models import User
from exporting.export import sanitize_csv
from permissions.models import AccessGrant
from projects.models import Project, SubProject
from tasks.models import Task


@pytest.fixture
def admin(db):
    return User.objects.create_superuser(email="a@example.com", name="Ada", password="pw-strong-123")


@pytest.fixture
def member(db):
    return User.objects.create_user(email="m@example.com", name="Mara", password="pw-strong-123")


@pytest.fixture
def sp(db):
    return SubProject.objects.create(project=Project.objects.create(name="Karuna"), name="Marketing")


def login(user):
    api = APIClient()
    res = api.post("/api/auth/login", {"email": user.email, "password": "pw-strong-123"}, format="json")
    api.credentials(HTTP_AUTHORIZATION=f"Bearer {res.data['access']}")
    return api


def read_csv(resp):
    return list(csv.reader(io.StringIO(resp.content.decode("utf-8"))))


# --- sanitization (unit) ----------------------------------------------------

@pytest.mark.parametrize("dangerous", ["=1+1", "+x", "-x", "@x", "\tx"])
def test_sanitize_prefixes_dangerous(dangerous):
    assert sanitize_csv(dangerous).startswith("'")


def test_sanitize_leaves_safe_values():
    assert sanitize_csv("Design flyer") == "Design flyer"
    assert sanitize_csv("") == ""


# --- CSV export -------------------------------------------------------------

def test_csv_has_header_and_rows(admin, sp):
    Task.objects.create(subproject=sp, title="Flyer")
    rows = read_csv(login(admin).get("/api/export?fmt=csv"))
    assert rows[0][0] == "ID"  # id is first, for round-trip import
    assert rows[0][1] == "Project"
    assert any("Flyer" in r for r in rows[1:])


def test_empty_export_still_valid(admin, db):
    rows = read_csv(login(admin).get("/api/export?fmt=csv"))
    assert rows == [[
        "ID", "Project", "Sub-project", "Title", "Status", "Priority", "Approval",
        "Deadline", "Start time", "End time", "Recurrence", "Assignees",
        "Details", "Requirements", "Links",
    ]]


def test_column_selection(admin, sp):
    Task.objects.create(subproject=sp, title="Flyer")
    rows = read_csv(login(admin).get("/api/export?fmt=csv&columns=title,priority"))
    assert rows[0] == ["Title", "Priority"]
    assert rows[1] == ["Flyer", "Medium"]


def test_filter_by_priority(admin, sp):
    Task.objects.create(subproject=sp, title="Hi", priority=5)
    Task.objects.create(subproject=sp, title="Lo", priority=1)
    rows = read_csv(login(admin).get("/api/export?fmt=csv&priority=5&columns=title"))
    flat = [c for r in rows for c in r]
    assert "Hi" in flat and "Lo" not in flat


def test_filter_unassigned(admin, sp):
    other = User.objects.create_user(email="x@example.com", name="X", password="pw-strong-123")
    t = Task.objects.create(subproject=sp, title="Assigned")
    t.assignees.add(other)
    Task.objects.create(subproject=sp, title="Free")
    rows = read_csv(login(admin).get("/api/export?fmt=csv&assignee=unassigned&columns=title"))
    flat = [c for r in rows for c in r]
    assert "Free" in flat and "Assigned" not in flat


def test_archived_excluded_by_default_included_on_flag(admin, sp):
    from django.utils import timezone
    Task.objects.create(subproject=sp, title="Archived", archived_at=timezone.now())
    default_rows = read_csv(login(admin).get("/api/export?fmt=csv&columns=title"))
    assert "Archived" not in [c for r in default_rows for c in r]
    incl = read_csv(login(admin).get("/api/export?fmt=csv&columns=title&archived=1"))
    assert "Archived" in [c for r in incl for c in r]


def test_formula_injection_sanitized_in_output(admin, sp):
    Task.objects.create(subproject=sp, title="=cmd|'/c calc'!A1")
    rows = read_csv(login(admin).get("/api/export?fmt=csv"))
    title_cell = rows[1][3]  # id, project, subproject, title
    assert title_cell.startswith("'=")  # neutralized


def test_commas_quotes_newlines_emoji_roundtrip(admin, sp):
    Task.objects.create(subproject=sp, title='A, "B"\nC 🎉')
    rows = read_csv(login(admin).get("/api/export?fmt=csv"))
    assert rows[1][3] == 'A, "B"\nC 🎉'  # csv quoting preserved it intact


def test_export_excludes_hidden_subprojects(member, sp):
    hidden = SubProject.objects.create(project=Project.objects.create(name="H"), name="Warehouse")
    AccessGrant.objects.create(user=member, subproject=sp, level="member")
    Task.objects.create(subproject=sp, title="Mine")
    Task.objects.create(subproject=hidden, title="NotMine")
    rows = read_csv(login(member).get("/api/export?fmt=csv"))
    flat = [c for r in rows for c in r]
    assert "Mine" in flat and "NotMine" not in flat


def test_export_excludes_pending(admin, sp):
    Task.objects.create(subproject=sp, title="Pending", approval_state="pending")
    rows = read_csv(login(admin).get("/api/export?fmt=csv"))
    flat = [c for r in rows for c in r]
    assert "Pending" not in flat


# --- XLSX export ------------------------------------------------------------

def test_xlsx_is_valid_workbook(admin, sp):
    from openpyxl import load_workbook

    Task.objects.create(subproject=sp, title="Flyer")
    resp = login(admin).get("/api/export?fmt=xlsx")
    assert resp["Content-Type"].startswith("application/vnd.openxmlformats")
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws["A1"].value == "ID"
    assert ws["B1"].value == "Project"
    assert any(cell.value == "Flyer" for col in ws.iter_cols() for cell in col)


def test_export_scope_multiple_subprojects(admin):
    p = Project.objects.create(name="MP")
    a = SubProject.objects.create(project=p, name="A")
    b = SubProject.objects.create(project=p, name="B")
    c = SubProject.objects.create(project=p, name="C")
    Task.objects.create(subproject=a, title="TA")
    Task.objects.create(subproject=b, title="TB")
    Task.objects.create(subproject=c, title="TC")
    rows = read_csv(login(admin).get(f"/api/export?fmt=csv&subprojects={a.id},{b.id}&columns=title"))
    flat = [c for r in rows for c in r]
    assert "TA" in flat and "TB" in flat and "TC" not in flat


def test_export_filter_by_group(admin):
    from accounts.models import Group
    p = Project.objects.create(name="GP")
    sp = SubProject.objects.create(project=p, name="S")
    g = Group.objects.create(name="Team")
    t = Task.objects.create(subproject=sp, title="GroupTask")
    t.assignee_groups.add(g)
    Task.objects.create(subproject=sp, title="NoGroup")
    rows = read_csv(login(admin).get(f"/api/export?fmt=csv&groups={g.id}&columns=title"))
    flat = [c for r in rows for c in r]
    assert "GroupTask" in flat and "NoGroup" not in flat

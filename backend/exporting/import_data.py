"""Task import: parse CSV / TSV / XLSX / JSON, preview (dry-run), then commit.

Two row kinds, told apart by the optional **Subtask** column:
  - Subtask cell BLANK  → a TASK row.
  - Subtask cell FILLED → a SUBTASK row: it adds the named subtask to the task in
    that same row (matched by ID or Title). Because a subtask is a mini-task, every
    other column (Assignees, Status, Priority, Deadline, …) applies to the SUBTASK.

Matching (tasks and a subtask's parent alike):
  - by the `id` column when present and valid;
  - otherwise by Title — 0 matches = create (tasks only), 1 = that task, N =
    "ambiguous": the preview lists the candidates and the user picks which to
    update / attach to (one / several / all) or create new.
  A subtask whose parent is a NEW task created by an earlier row in the same sheet
  attaches to it (rows are processed top-to-bottom).

Updates are PARTIAL and safe: only columns that carry a value overwrite; a blank
or absent column leaves the existing field untouched, so a sparse sheet never
wipes assignees / deadline / status. Creates fill blanks with defaults. Subtasks
are added-only and de-duped by title (re-importing never doubles them).

Everything is scoped to the importing admin's organization. A restore-point
checkpoint is saved before any committing import (see views) so a bad import is
one click to undo.
"""

import csv
import io
import json
import re
from datetime import datetime

from django.db import transaction

from accounts.models import User
from projects.models import Project, SubProject
from tasks.models import Status, Subtask, Task

from .export import COLUMNS

MAX_ROWS = 5000
# columns we don't import (server-controlled or too complex to round-trip safely)
SKIP_COLUMNS = {"approval", "recurrence"}
# extra headers the importer understands that aren't export COLUMNS
_EXTRA_HEADERS = {"subtask": "subtask", "sub-task": "subtask",
                  "subtask title": "subtask", "sub-task title": "subtask"}


# ── header mapping ──────────────────────────────────────────────────────────

def _col_key_for(header):
    """Map an incoming header (label OR key, any case) to a column key."""
    h = (header or "").strip().lower()
    for key, (label, _) in COLUMNS.items():
        if h == key or h == label.lower():
            return key
    return _EXTRA_HEADERS.get(h)


# ── parsing ─────────────────────────────────────────────────────────────────

def parse(fmt, content):
    """Return a list of {col_key: str_value} dicts. `content` is text for
    csv/tsv/json, bytes for xlsx. Raises ValueError on malformed input."""
    fmt = (fmt or "csv").lower()
    if fmt == "json":
        return _parse_json(content)
    if fmt == "xlsx":
        return _parse_xlsx(content)
    return _parse_delimited(content, "\t" if fmt == "tsv" else _sniff_delim(content))


def _sniff_delim(text):
    first = (text or "").splitlines()[0] if text else ""
    return "\t" if ("\t" in first and first.count("\t") >= first.count(",")) else ","


def _rows_from_matrix(matrix):
    if not matrix:
        return []
    header_keys = [_col_key_for(h) for h in matrix[0]]
    out = []
    for raw in matrix[1:]:
        if not any((c or "").strip() for c in raw):
            continue  # skip blank lines
        row = {}
        for key, val in zip(header_keys, raw):
            if key and key not in SKIP_COLUMNS:
                row[key] = "" if val is None else str(val).strip()
        out.append(row)
    return out


def _parse_delimited(text, delimiter):
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return _rows_from_matrix(list(reader))


def _parse_xlsx(content):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    matrix = [[c if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)]
    return _rows_from_matrix(matrix)


def _parse_json(text):
    data = json.loads(text)
    if not isinstance(data, list):
        raise ValueError("JSON import must be a list of task objects.")
    out = []
    for obj in data:
        if not isinstance(obj, dict):
            raise ValueError("Each JSON item must be an object.")
        row = {}
        for k, v in obj.items():
            key = _col_key_for(k)
            if key and key not in SKIP_COLUMNS:
                row[key] = "" if v is None else str(v).strip()
        out.append(row)
    return out


# ── value coercion ──────────────────────────────────────────────────────────

def _parse_date(s):
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"bad date '{s}' (use YYYY-MM-DD)")


def _parse_time(s):
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    raise ValueError(f"bad time '{s}' (use HH:MM)")


_PRIORITY_LABELS = {"lowest": 1, "low": 2, "medium": 3, "high": 4, "highest": 5}


def _parse_priority(s):
    s = s.lower()
    if s.isdigit() and 1 <= int(s) <= 5:
        return int(s)
    if s in _PRIORITY_LABELS:
        return _PRIORITY_LABELS[s]
    raise ValueError(f"bad priority '{s}'")


def _parse_links(s):
    return [p for p in (s or "").replace(",", " ").split() if p]


def _status_maps():
    by = {}
    for st in Status.objects.all():
        by[st.key.lower()] = st.key
        by[st.label.lower()] = st.key
    initial = (
        Status.objects.filter(is_initial=True).values_list("key", flat=True).first()
        or Status.objects.order_by("order", "id").values_list("key", flat=True).first()
        or "todo"
    )
    return by, initial


def _user_maps(org):
    by_email, by_name = {}, {}
    qs = User.objects.filter(is_active=True)
    if org is not None:
        qs = qs.filter(memberships__organization=org, memberships__is_active=True).distinct()
    for u in qs:
        by_email[u.email.lower()] = u.id
        if u.name:
            by_name.setdefault(u.name.lower(), u.id)
    return by_email, by_name


# ── resolution (pure: no DB writes) ───────────────────────────────────────────

class Ctx:
    """Lookups the resolver needs, built once per preview/commit. Scoped to `org`
    so name/project matching never crosses the tenant line."""

    def __init__(self, org=None):
        self.org = org
        proj_qs = Project.objects.all()
        sub_qs = SubProject.objects.select_related("project").all()
        task_qs = Task.objects.filter(archived_at__isnull=True).select_related("subproject__project")
        if org is not None:
            proj_qs = proj_qs.filter(organization=org)
            sub_qs = sub_qs.filter(project__organization=org)
            task_qs = task_qs.filter(subproject__project__organization=org)

        self.projects = {p.name.lower(): p for p in proj_qs}
        self.subs = {(sp.project.name.lower(), sp.name.lower()): sp for sp in sub_qs}
        self.task_ids = set()
        self.tasks_by_title = {}
        for tk in task_qs:
            self.task_ids.add(tk.id)
            self.tasks_by_title.setdefault(tk.title.lower(), []).append(
                {"id": tk.id, "title": tk.title,
                 "project": tk.subproject.project.name, "subproject": tk.subproject.name}
            )
        self.status_map, self.initial_status = _status_maps()
        self.email_map, self.name_map = _user_maps(org)


def _provided_fields(row, r, ctx):
    """Scalar fields parsed from the row — only columns that carry a value land
    here; blanks/absent are omitted so an update leaves them untouched. Shared by
    task rows and subtask rows (a subtask has the same fields)."""
    provided = {}

    def take(key, parser):
        raw = (row.get(key) or "").strip()
        if raw == "":
            return
        try:
            provided[key] = parser(raw)
        except ValueError as e:
            r["errors"].append(str(e))

    take("details", lambda s: s)
    take("requirements", lambda s: s)
    take("links", _parse_links)
    take("deadline", _parse_date)
    take("priority", _parse_priority)

    raw_status = (row.get("status") or "").strip()
    if raw_status:
        key = ctx.status_map.get(raw_status.lower())
        if key:
            provided["status"] = key
        else:
            r["warnings"].append(f"unknown status '{raw_status}' — left unchanged")

    st_raw = (row.get("start_time") or "").strip()
    et_raw = (row.get("end_time") or "").strip()
    if st_raw or et_raw:
        if not (st_raw and et_raw):
            r["errors"].append("set both start and end time, or neither")
        else:
            try:
                st, et = _parse_time(st_raw), _parse_time(et_raw)
                if et <= st:
                    r["errors"].append("end time must be after start time")
                else:
                    provided["start_time"], provided["end_time"] = st, et
            except ValueError as e:
                r["errors"].append(str(e))
    return provided


def _match_parent(row, ctx, pending_titles):
    """Resolve a row to an existing task: (match_id, candidates, parent_pending).
    match_id = a single existing task; candidates = >1 same-named tasks (ambiguous);
    parent_pending = the title belongs to a NEW task an earlier row will create."""
    title = (row.get("title") or "").strip()
    raw_id = (row.get("id") or "").strip()
    warnings = []
    if raw_id:
        if raw_id.isdigit() and int(raw_id) in ctx.task_ids:
            return int(raw_id), [], False, warnings
        if raw_id.isdigit():
            warnings.append(f"ID {raw_id} not found — matching by name instead")
        else:
            warnings.append(f"ignoring non-numeric ID '{raw_id}'")
    if title:
        cands = ctx.tasks_by_title.get(title.lower(), [])
        if len(cands) == 1:
            return cands[0]["id"], [], False, warnings
        if len(cands) > 1:
            return None, cands, False, warnings
        if title.lower() in pending_titles:
            return None, [], True, warnings
    return None, [], False, warnings


def _resolve_row(row, ctx, pending_titles):
    """Compute a row's resolution (no DB writes)."""
    r = {"errors": [], "warnings": [], "provided": {}, "assignee_ids": None,
         "match_id": None, "candidates": [], "parent_pending": False,
         "is_subtask": False, "subtask": "",
         "will_create_project": False, "will_create_subproject": False}

    title = (row.get("title") or "").strip()
    r["title"] = title
    project_name = (row.get("project") or "").strip()
    sub_raw = (row.get("subproject") or "").strip()
    r["project"], r["subproject"] = project_name, (sub_raw or "General")
    subtask_title = (row.get("subtask") or "").strip()
    r["is_subtask"] = bool(subtask_title)
    r["subtask"] = subtask_title

    match_id, candidates, parent_pending, warnings = _match_parent(row, ctx, pending_titles)
    r["match_id"], r["candidates"], r["parent_pending"] = match_id, candidates, parent_pending
    r["warnings"].extend(warnings)

    if r["is_subtask"]:
        # A subtask needs an existing/pending parent — it never creates a task.
        if not (match_id or candidates or parent_pending):
            who = f"'{title}'" if title else "(no Title/ID given)"
            r["errors"].append(f"no task {who} found to add subtask '{subtask_title}' to")
        r["action"] = "ambiguous" if candidates else "subtask"
    else:
        r["action"] = "ambiguous" if candidates else ("update" if match_id else "create")
        if r["action"] == "create":
            if not title:
                r["errors"].append("Title is required")
            if not project_name:
                r["errors"].append("Project is required")
            else:
                sub_name = sub_raw or "General"
                if project_name.lower() not in ctx.projects:
                    r["will_create_project"] = True
                    r["will_create_subproject"] = sub_name.lower() != "general"
                elif (project_name.lower(), sub_name.lower()) not in ctx.subs:
                    r["will_create_subproject"] = True

    r["provided"] = _provided_fields(row, r, ctx)

    raw_assignees = (row.get("assignees") or "").strip()
    if raw_assignees:
        ids = []
        for part in [p.strip() for p in raw_assignees.replace(";", ",").split(",") if p.strip()]:
            uid = ctx.email_map.get(part.lower()) or ctx.name_map.get(part.lower())
            if uid:
                ids.append(uid)
            else:
                r["warnings"].append(f"no user matched '{part}' — skipped")
        r["assignee_ids"] = ids
    return r


# ── preview & commit ───────────────────────────────────────────────────────

def preview(rows, org=None):
    if len(rows) > MAX_ROWS:
        raise ValueError(f"Too many rows ({len(rows)}); max {MAX_ROWS}.")
    ctx = Ctx(org)
    pending = set()   # titles of CREATE task rows seen so far (subtasks may attach)
    out, new_projects, new_subs = [], set(), set()
    counts = {"create": 0, "update": 0, "subtask": 0, "ambiguous": 0, "error": 0}
    for i, row in enumerate(rows):
        res = _resolve_row(row, ctx, pending)
        if not res["is_subtask"] and res["action"] == "create" and res["title"] and not res["errors"]:
            pending.add(res["title"].lower())
        if res["will_create_project"]:
            new_projects.add(res["project"])
        if res["will_create_subproject"]:
            new_subs.add(f"{res['project']} / {res['subproject']}")
        action = "error" if res["errors"] else res["action"]
        counts[action] = counts.get(action, 0) + 1
        out.append({
            "row": i, "action": action, "is_subtask": res["is_subtask"], "subtask": res["subtask"],
            "match_id": res["match_id"], "candidates": res["candidates"],
            "parent_pending": res["parent_pending"], "title": res["title"],
            "project": res["project"], "subproject": res["subproject"],
            "will_create_project": res["will_create_project"],
            "will_create_subproject": res["will_create_subproject"],
            "update_fields": sorted(res["provided"].keys()) + (["assignees"] if res["assignee_ids"] is not None else []),
            "errors": res["errors"], "warnings": res["warnings"],
        })
    return {
        "rows": out,
        "new_projects": sorted(new_projects),
        "new_subprojects": sorted(new_subs),
        "summary": counts,
        "total": len(rows),
    }


def _apply_task_update(task, res):
    """Apply a row's partial fields + assignees to an existing task. A non-blank
    Title renames the task (for a name-matched row it just equals the current one)."""
    changed = []
    if res["title"] and res["title"] != task.title:
        task.title = res["title"]
        changed.append("title")
    for k, v in res["provided"].items():
        setattr(task, k, v)
        changed.append(k)
    if changed:
        task.save(update_fields=changed)
    if res["assignee_ids"] is not None:
        task.assignees.set(res["assignee_ids"])


def _add_subtask(task, res):
    """Add one subtask (deduped by title) to `task`, with the row's fields. Returns
    1 if created, 0 if a same-titled subtask already existed."""
    title = res["subtask"]
    if any(s.title.lower() == title.lower() for s in task.subtasks.all()):
        return 0
    f = res["provided"]
    st = Subtask(
        task=task, title=title,
        status=f.get("status", "todo"), priority=f.get("priority", 3),
        details=f.get("details", ""), requirements=f.get("requirements", ""),
        links=f.get("links", []), deadline=f.get("deadline"),
        start_time=f.get("start_time"), end_time=f.get("end_time"),
    )
    st.save()
    if res["assignee_ids"]:
        st.assignees.set(res["assignee_ids"])
    return 1


@transaction.atomic
def commit(user, rows, decisions, org=None):
    """Apply the import. `decisions` maps str(row_index) -> a decision:
      "skip"               — do nothing for this row
      "create"             — force-create a new task (even if it matched a name)
      "overwrite"          — update the matched task (the default 1:1 match)
      {"ids": [int, ...]}  — target exactly these tasks (an ambiguous row's pick)
    A missing entry follows the row's natural action; an ambiguous row with no
    decision is skipped (the user must choose). Rows with errors are skipped."""
    if len(rows) > MAX_ROWS:
        raise ValueError(f"Too many rows ({len(rows)}); max {MAX_ROWS}.")
    ctx = Ctx(org)
    proj_cache = dict(ctx.projects)
    sub_cache = dict(ctx.subs)
    created_titles = {}   # title.lower -> [Task] created earlier in this run
    pending = set()
    result = {"created": 0, "updated": 0, "subtasks": 0, "skipped": 0, "errors": 0}

    def get_project(name):
        p = proj_cache.get(name.lower())
        if p is None:
            p = Project.objects.create(name=name, **({"organization": org} if org is not None else {}))
            proj_cache[name.lower()] = p
            for sp in p.subprojects.all():
                sub_cache[(name.lower(), sp.name.lower())] = sp
        return p

    def get_sub(pname, sname):
        key = (pname.lower(), sname.lower())
        sp = sub_cache.get(key)
        if sp is None:
            sp = SubProject.objects.create(project=get_project(pname), name=sname)
            sub_cache[key] = sp
        return sp

    valid_ids = set(ctx.task_ids)

    for i, row in enumerate(rows):
        # pending grows as we go (like preview) so a subtask only attaches to a NEW
        # task whose row came ABOVE it; created_titles holds the real saved tasks.
        res = _resolve_row(row, ctx, pending)
        decision = decisions.get(str(i))
        if res["errors"]:
            result["errors"] += 1
            continue
        if decision == "skip":
            result["skipped"] += 1
            continue

        # which existing task ids does this row target (vs. create-new)?
        target_ids = None
        if decision == "create" and not res["is_subtask"]:
            target_ids = []
        elif isinstance(decision, dict) and isinstance(decision.get("ids"), list):
            # The user picked specific tasks. Coerce once + tolerate junk (a malformed
            # client could send non-numerics); only ids visible in THIS org survive
            # (valid_ids is org-scoped). If NONE survive, skip — never silently create.
            chosen = []
            for x in decision["ids"]:
                try:
                    tid = int(x)
                except (ValueError, TypeError):
                    continue
                if tid in valid_ids:
                    chosen.append(tid)
            if not chosen:
                result["skipped"] += 1
                continue
            target_ids = chosen
        elif res["match_id"]:
            target_ids = [res["match_id"]]
        elif res["candidates"]:                 # ambiguous, no pick → leave untouched
            result["skipped"] += 1
            continue
        elif res["parent_pending"]:             # subtask attaching to a new task above
            target_ids = "pending"

        if res["is_subtask"]:
            parents = (created_titles.get(res["title"].lower(), []) if target_ids == "pending"
                       else [Task.objects.filter(pk=t).first() for t in (target_ids or [])])
            added = False
            for task in parents:
                if task is not None:
                    result["subtasks"] += _add_subtask(task, res)
                    added = True
            if not added:
                result["skipped"] += 1
            continue

        if target_ids:   # update existing task(s)
            for tid in target_ids:
                task = Task.objects.filter(pk=tid).first()
                if task is not None:
                    _apply_task_update(task, res)
                    result["updated"] += 1
        else:            # create a new task
            sub = get_sub(res["project"], res["subproject"] or "General")
            f = res["provided"]
            task = Task(
                subproject=sub, title=res["title"], created_by=user,
                approval_state=Task.Approval.APPROVED,
                priority=f.get("priority", 3), status=f.get("status", ctx.initial_status),
                details=f.get("details", ""), requirements=f.get("requirements", ""),
                links=f.get("links", []), deadline=f.get("deadline"),
                start_time=f.get("start_time"), end_time=f.get("end_time"),
            )
            task.save()
            if res["assignee_ids"]:
                task.assignees.set(res["assignee_ids"])
            created_titles.setdefault(task.title.lower(), []).append(task)
            pending.add(task.title.lower())   # later subtask rows can attach to it
            result["created"] += 1

    return result
